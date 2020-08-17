import win32com.client
import openpyxl
import time
import re
import uuid
import datetime as dt
from pathlib import Path

def main():

    # Create workbook objects
    # wb_input = openpyxl.load_workbook(r"\\fssg52\FSP22\Business Analytics Working\Dev\Jerome\MasterTrackerSheet.xlsx")
    wb_input = openpyxl.load_workbook(Path(r"//fssg52/FSP22/Business Analytics Working/Dev/Jerome/") / "MasterTrackerSheet.xlsx")

    # Specifying sheets to work on
    input_sheet = wb_input["Tracker"]

    all_tasks_list = [cell.value for cell in input_sheet["B"] if cell.value is not None]

    # Create outlook objects
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    accounts= win32com.client.DispatchEx("Outlook.Application").Session.Accounts
    interval = 1

    # this is set to one hour ago
    lastHourDateTime = dt.datetime.now() - dt.timedelta(hours = interval)

    # Scans inbox for appropriate email addresses
    inbox = outlook.GetDefaultFolder(6)

    # retrieve all emails in the inbox, then sort them from most recently received to oldest (False will give you the reverse). Not strictly necessary, but good to know if order matters for your search
    messages = inbox.Items
    messages.Sort("[ReceivedTime]", True)

    # restrict to messages from the past hour based on ReceivedTime using the dates defined above.
    # lastHourMessages will contain only emails with a ReceivedTime later than an hour ago
    # The way the datetime is formatted DOES matter (must match with your system's time format); You can't add seconds here.
    lastHourMessages = messages.Restrict("[ReceivedTime] >= '" +lastHourDateTime.strftime('%d/%m/%Y %H:%M %p')+"'")

    print("Current time: "+ dt.datetime.now().strftime('%d/%m/%Y %H:%M %p'))
    print(f"Message(s) from the past {interval} hour(s):")

    for message in lastHourMessages:
        print(message.Subject)
        
        if message.Class == 43:
            if message.SenderEmailType == "EX":
                print(message.Sender.GetExchangeUser().PrimarySmtpAddress)
            else:
                print(message.SenderEmailAddress)
        print()

    # Checks if appropriate email addresses have sent requests via template (with correct subject too)
    valid_email_ext = "@ocbc.com"
    valid_email_ext2 = "@ocbc.local"
    valid_subject = "Campaign Request"

    email_ext_regex = re.compile(rf".+{valid_email_ext}|{valid_email_ext2}", re.I)
    subject_regex = re.compile(rf"^{valid_subject}.+", re.I)

    for message in lastHourMessages:
        
        if message.Class == 43:
            if message.SenderEmailType == "EX": 
                email_address = message.Sender.GetExchangeUser().PrimarySmtpAddress
            else:
                email_address = message.SenderEmailAddress

            if email_ext_regex.search(email_address) and subject_regex.search(message.Subject):
                message_body = message.Body.replace("\r\n", " ")
                message_list = message_body.split(" ")
                message_list = [item for item in message_list if item]  # removes blank
                valid_message = []

                for word in message_list:
                    if word.strip() == "BU:" or word.strip() == "DueDate:":  # Searches the body for these keywords (change this if you want to add to template)
                        valid_message.append(word)

                # Checks if email is valid
                if len(valid_message) == 2:
                        print(f"Valid message(s) from the past {interval} hour(s):")
                        print(message.Subject)

                        # Saves all attachments to a central repo
                        # First generate a list of uuids to uniquely identify attachments (as attachment names may be same across email)
                        id_list = []

                        for attachment in message.Attachments:
                            unique_identifier = uuid.uuid1()
                            attachment.SaveAsFile(Path(r"//fssg52/FSP22/Business Analytics Working/Dev/Jerome/") / f"{unique_identifier} {attachment.FileName}")
                            id_list.append(str(unique_identifier))
                            print(f"{attachment} saved.")

                        # Update MasterTrackerSheet with valid email content
                        row = 1
                        while input_sheet[f"A{row}"].value is not None:  # Iterate until an empty row is reached
                            row += 1

                        # Inserts only if same task isn't found in tracking sheet
                        if message.Subject not in all_tasks_list:
                            task_list = [row - 1, 
                                        message.Subject, 
                                        "", 
                                        "", 
                                        ", ".join([attachment.FileName for attachment in message.Attachments]), 
                                        ", ".join(id_list),
                                        message_list[message_list.index("BU:") + 1], 
                                        email_address, 
                                        message_list[message_list.index("DueDate:") + 1], 
                                        message.ReceivedTime, "", "BACKLOG"]

                            for col, val in enumerate(task_list, start=1):
                                input_sheet.cell(row=row, column=col).value = val

                            print("Appended to master tracker sheet.")

                        # Relocate email to another folder
                        message.Move(outlook.GetDefaultFolder(6).Folders("Campaign Requests"))
                        print("Moved email to Campaign Requests folder.")
                  

    # Updates tracker sheet with latest update date based on reply emails
    reply_regex = re.compile(r"^RE: .+", re.I)  # Reply emails start with RE: ...

    for message in lastHourMessages:
        if reply_regex.search(message.Subject):
            if message.Subject[4:] in all_tasks_list:  # If reply email corresponds with a task in master tracker
                row_number = all_tasks_list.index(message.Subject[4:]) + 2
                input_sheet.cell(row=row_number, column=8).value = message.ReceivedTime
                break  # exits for loop once first reply email matches (to ensure latest update time is extracted)

    # Forwards email to assignee as defined in master tracker (only if it hasn't sent it before)
    campaign_request_folder = outlook.GetDefaultFolder(6).Folders("Campaign Requests")
    campaign_request_list = campaign_request_folder.Items
    for campaign in campaign_request_list:
        try:
            row_number = all_tasks_list.index(campaign.Subject) + 2

            if input_sheet[f"C{row_number}"].value and input_sheet[f"D{row_number}"].value != "Y":  # If email hasn't been forwarded yet and assignee has been added
                new_msg = campaign.Forward()
                new_msg.Body = "[This task has been assigned to you.]\n\n" + campaign.Body
                new_msg.Subject = "[Assigned] " + campaign.Subject
                new_msg.To = input_sheet[f"C{row_number}"].value
                new_msg.Send()
                input_sheet[f"D{row_number}"].value = "Y"
                print(f"Assigned and forwarded email {new_msg.Subject}.")
        except:
            print("Not forwarded")
            
    wb_input.save(Path(r"//fssg52/FSP22/Business Analytics Working/Dev/Jerome/") / "MasterTrackerSheet.xlsx")


if __name__ == '__main__':
    main()

